import pandas as pd
import requests as req
import xlsxwriter
import xlrd
import openpyxl
import json
from json2xml import json2xml
import warnings

wb = openpyxl.load_workbook('urlsexcel.xlsx')
sheet1 = wb.get_sheet_by_name('staging')
sheet2 = wb.get_sheet_by_name('apis')

#function to handle post requests
def postRequestMethod(sheet,ind,req):
    URL = sheet.cell(row=ind,column=1).value
    print(URL)
    Pram = json.loads(sheet.cell(row=ind,column=3).value)
    print(Pram)
    Head = json.loads(sheet.cell(row=ind,column=4).value)
    print(Head)
    response = req.post(url=URL,data=Pram,headers=Head)
    print(response.json())
    sheet.cell(row=ind,column=7).value=str(response.json())
    sheet.cell(row=ind,column=6).value=str(response.headers)
    sheet.cell(row=ind,column=5).value=response.status_code
    assert response.status_code == 200
def getRequestMethod(sheet,ind,req):
    URL = sheet.cell(row=ind,column=1).value
    print(URL)
    Pram = json.loads(sheet.cell(row=ind,column=3).value)
    print(Pram)
    Head = json.loads(sheet.cell(row=ind,column=4).value)
    print(Head)
    response = req.get(url=URL,data=Pram,headers=Head)
    print(response.json())
    sheet.cell(row=ind,column=7).value=str(response.json())
    sheet.cell(row=ind,column=6).value=str(response.headers)
    sheet.cell(row=ind,column=5).value=response.status_code
    assert response.status_code == 200
   # print(json2xml.Json2xml(response.json()).to_xml())

#for loop to handle web url healthcheck 
for ind in range(2,sheet1.max_row+1):
    resp = req.get(sheet1.cell(row=ind,column=1).value)
    print(resp.status_code)
    print(resp.raise_for_status)
    print(resp.headers)
    print(resp.json)
    sheet1.cell(row=ind,column=2).value = resp.status_code
    if(resp.status_code!=200):
        warnings.warn(str(resp.status_code))

#for loop to go over rest apis
for ind in range(2,sheet2.max_row+1):
    if(sheet2.cell(row=ind, column=2).value=='Post'):
        postRequestMethod(sheet2,ind,req)
    else:
        getRequestMethod(sheet2,ind,req)
wb.save('urlsexcel.xlsx')
   













































































